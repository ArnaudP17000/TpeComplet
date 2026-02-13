"""
Module de Gestion des Terminaux de Paiement Électronique (T.P.E.)
Auteur: Spécialiste Python
Date: 2026-02-13 - Version 1.5 - Numéro de série TPE
"""

import json
import pickle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dataclasses import dataclass, asdict
from typing import List, Optional
from datetime import datetime
import re
from pathlib import Path


@dataclass
class Regisseur:
    """Classe pour gérer les informations du régisseur"""
    prenom: str
    nom: str
    telephone: str
    
    def __str__(self):
        return f"{self.prenom} {self.nom} - {self.telephone}"
    
    def to_dict(self):
        return {
            'prenom': self.prenom,
            'nom': self.nom,
            'telephone': self.telephone
        }
    
    @classmethod
    def from_dict(cls, data):
        return cls(**data)


@dataclass
class ConfigurationReseau:
    """Classe pour la configuration réseau (si Ethernet sélectionné)"""
    adresse_ip: str
    masque: str
    passerelle: str
    
    def __post_init__(self):
        # Validation des adresses IP
        self._valider_ip(self.adresse_ip, "Adresse IP")
        self._valider_ip(self.masque, "Masque")
        self._valider_ip(self.passerelle, "Passerelle")
    
    def _valider_ip(self, ip: str, nom: str):
        """Valide le format d'une adresse IP"""
        pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
        if not re.match(pattern, ip):
            raise ValueError(f"{nom} invalide: {ip}")
        
        octets = ip.split('.')
        for octet in octets:
            if int(octet) > 255:
                raise ValueError(f"{nom} invalide: octet > 255")
    
    def to_dict(self):
        return {
            'adresse_ip': self.adresse_ip,
            'masque': self.masque,
            'passerelle': self.passerelle
        }
    
    @classmethod
    def from_dict(cls, data):
        if data is None:
            return None
        return cls(**data)


@dataclass
class AccesBackoffice:
    """Classe pour gérer l'accès backoffice"""
    actif: bool
    email: Optional[str] = None
    
    def __post_init__(self):
        if self.actif and self.email:
            self._valider_email()
    
    def _valider_email(self):
        """Valide le format de l'email"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(pattern, self.email):
            raise ValueError(f"Email invalide: {self.email}")
    
    def to_dict(self):
        return {
            'actif': self.actif,
            'email': self.email
        }
    
    @classmethod
    def from_dict(cls, data):
        return cls(**data)


@dataclass
class TypeTPE:
    """Classe pour gérer le type de TPE"""
    ethernet: bool = False
    quatre_cinq_g: bool = False
    config_reseau: Optional[ConfigurationReseau] = None
    
    def __post_init__(self):
        if self.ethernet and self.config_reseau is None:
            raise ValueError("Configuration réseau requise pour le type Ethernet")
    
    def to_dict(self):
        return {
            'ethernet': self.ethernet,
            'quatre_cinq_g': self.quatre_cinq_g,
            'config_reseau': self.config_reseau.to_dict() if self.config_reseau else None
        }
    
    @classmethod
    def from_dict(cls, data):
        config = ConfigurationReseau.from_dict(data.get('config_reseau')) if data.get('config_reseau') else None
        return cls(
            ethernet=data['ethernet'],
            quatre_cinq_g=data['quatre_cinq_g'],
            config_reseau=config
        )


@dataclass
class CarteCommercant:
    """Classe pour gérer une carte commerçant avec son numéro de série TPE"""
    numero: str  # Numéro de carte (alphanumérique)
    numero_serie_tpe: Optional[str] = None  # Numéro de série du TPE (optionnel)
    
    def to_dict(self):
        return {
            'numero': self.numero,
            'numero_serie_tpe': self.numero_serie_tpe
        }
    
    @classmethod
    def from_dict(cls, data):
        # Si c'est juste un string ou int (ancienne version), le convertir
        if isinstance(data, str) or isinstance(data, int):
            return cls(numero=str(data), numero_serie_tpe=None)
        # Sinon c'est un dict complet
        return cls(
            numero=str(data.get('numero', '')),
            numero_serie_tpe=data.get('numero_serie_tpe')
        )


@dataclass
class TPE:
    """Classe principale représentant un Terminal de Paiement Électronique"""
    service: str
    regisseur: Regisseur
    regisseurs_suppleants: str
    cartes_commercant: List[CarteCommercant]  # Liste de CarteCommercant
    shop_id: int
    acces_backoffice: AccesBackoffice
    modele_tpe: str
    type_tpe: TypeTPE
    nombre_tpe: int = 1
    date_creation: str = None
    
    def __post_init__(self):
        if self.date_creation is None:
            self.date_creation = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Validation des cartes commerçant
        if not self.cartes_commercant or len(self.cartes_commercant) == 0:
            raise ValueError("Au moins une carte commerçant est requise")
        
        for carte in self.cartes_commercant:
            if not isinstance(carte, CarteCommercant):
                raise ValueError(f"Carte commerçant invalide: {carte}")
            if not carte.numero or not carte.numero.strip():
                raise ValueError(f"Numéro de carte invalide")
            if len(carte.numero) > 50:
                raise ValueError(f"Numéro de carte trop long (max 50 caractères)")
            # Numéro de série optionnel
            if carte.numero_serie_tpe and len(carte.numero_serie_tpe) > 100:
                raise ValueError(f"Numéro de série TPE trop long (max 100 caractères)")
        
        if not isinstance(self.shop_id, int) or self.shop_id < 0:
            raise ValueError("ShopID doit être un nombre positif")
        if not isinstance(self.nombre_tpe, int) or self.nombre_tpe < 1:
            raise ValueError("Le nombre de TPE doit être au minimum 1")
    
    def to_dict(self):
        """Convertit le TPE en dictionnaire"""
        return {
            'service': self.service,
            'regisseur': self.regisseur.to_dict(),
            'regisseurs_suppleants': self.regisseurs_suppleants,
            'cartes_commercant': [carte.to_dict() for carte in self.cartes_commercant],
            'shop_id': self.shop_id,
            'acces_backoffice': self.acces_backoffice.to_dict(),
            'modele_tpe': self.modele_tpe,
            'type_tpe': self.type_tpe.to_dict(),
            'nombre_tpe': self.nombre_tpe,
            'date_creation': self.date_creation
        }
    
    @classmethod
    def from_dict(cls, data):
        """Crée un TPE à partir d'un dictionnaire"""
        # Compatibilité avec anciennes versions
        cartes_data = data.get('cartes_commercant')
        if cartes_data is None:
            # Ancienne version avec carte_commercant unique
            carte_unique = data.get('carte_commercant')
            if carte_unique is not None:
                cartes_data = [carte_unique]
            else:
                cartes_data = []
        
        # Convertir en objets CarteCommercant
        cartes = [CarteCommercant.from_dict(c) for c in cartes_data]
        
        return cls(
            service=data['service'],
            regisseur=Regisseur.from_dict(data['regisseur']),
            regisseurs_suppleants=data['regisseurs_suppleants'],
            cartes_commercant=cartes,
            shop_id=data['shop_id'],
            acces_backoffice=AccesBackoffice.from_dict(data['acces_backoffice']),
            modele_tpe=data['modele_tpe'],
            type_tpe=TypeTPE.from_dict(data['type_tpe']),
            nombre_tpe=data.get('nombre_tpe', 1),
            date_creation=data.get('date_creation')
        )


class GestionnaireTPE:
    """Gestionnaire principal pour la gestion des TPE"""
    
    def __init__(self):
        self.tpes: List[TPE] = []
        self.fichier_sauvegarde = "tpe_data.pkl"
        self.fichier_backup = "tpe_backup.json"
    
    def ajouter_tpe(self, tpe: TPE) -> bool:
        """Ajoute un nouveau TPE"""
        try:
            # Si ShopID est 0, générer automatiquement
            if tpe.shop_id == 0:
                if self.tpes:
                    max_shop_id = max(t.shop_id for t in self.tpes)
                    tpe.shop_id = max_shop_id + 1
                else:
                    tpe.shop_id = 1
            
            # Vérification unicité ShopID
            if any(t.shop_id == tpe.shop_id for t in self.tpes):
                raise ValueError(f"ShopID {tpe.shop_id} existe déjà")
            
            self.tpes.append(tpe)
            return True
        except Exception as e:
            return False
    
    def supprimer_tpe(self, shop_id: int) -> bool:
        """Supprime un TPE par son ShopID"""
        try:
            self.tpes = [t for t in self.tpes if t.shop_id != shop_id]
            return True
        except Exception as e:
            return False
    
    def rechercher_tpe(self, shop_id: int) -> Optional[TPE]:
        """Recherche un TPE par son ShopID"""
        for tpe in self.tpes:
            if tpe.shop_id == shop_id:
                return tpe
        return None
    
    def modifier_tpe(self, shop_id: int, nouveau_tpe: TPE) -> bool:
        """Modifie un TPE existant"""
        try:
            for i, tpe in enumerate(self.tpes):
                if tpe.shop_id == shop_id:
                    nouveau_tpe.date_creation = tpe.date_creation
                    self.tpes[i] = nouveau_tpe
                    return True
            return False
        except Exception as e:
            return False
    
    def exporter_excel(self, nom_fichier: str = "tpe_export.xlsx") -> bool:
        """
        Exporte la liste des TPE au format Excel (.xlsx)
        Retourne True si succès, False sinon
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gestion TPE"
            
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            headers = [
                "Service", "Régisseur Prénom", "Régisseur Nom", "Régisseur Téléphone",
                "Régisseurs Suppléants", "Cartes Commerçant", "Numéros Série TPE", "ShopID", "Nombre de TPE",
                "Accès Backoffice", "Email Backoffice", "Modèle TPE",
                "Type Ethernet", "Type 4/5G", "Adresse IP", "Masque", "Passerelle",
                "Date Création"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            for row_idx, tpe in enumerate(self.tpes, start=2):
                ws.cell(row=row_idx, column=1, value=tpe.service)
                ws.cell(row=row_idx, column=2, value=tpe.regisseur.prenom)
                ws.cell(row=row_idx, column=3, value=tpe.regisseur.nom)
                ws.cell(row=row_idx, column=4, value=tpe.regisseur.telephone)
                ws.cell(row=row_idx, column=5, value=tpe.regisseurs_suppleants)
                
                # Cartes commerçant avec numéros de série
                cartes_str = ", ".join([c.numero for c in tpe.cartes_commercant])
                ws.cell(row=row_idx, column=6, value=cartes_str)
                
                numeros_serie_str = ", ".join([c.numero_serie_tpe or "N/A" for c in tpe.cartes_commercant])
                ws.cell(row=row_idx, column=7, value=numeros_serie_str)
                
                ws.cell(row=row_idx, column=8, value=tpe.shop_id)
                ws.cell(row=row_idx, column=9, value=tpe.nombre_tpe)
                ws.cell(row=row_idx, column=10, value="Oui" if tpe.acces_backoffice.actif else "Non")
                ws.cell(row=row_idx, column=11, value=tpe.acces_backoffice.email or "")
                ws.cell(row=row_idx, column=12, value=tpe.modele_tpe)
                ws.cell(row=row_idx, column=13, value="Oui" if tpe.type_tpe.ethernet else "Non")
                ws.cell(row=row_idx, column=14, value="Oui" if tpe.type_tpe.quatre_cinq_g else "Non")
                
                if tpe.type_tpe.config_reseau:
                    ws.cell(row=row_idx, column=15, value=tpe.type_tpe.config_reseau.adresse_ip)
                    ws.cell(row=row_idx, column=16, value=tpe.type_tpe.config_reseau.masque)
                    ws.cell(row=row_idx, column=17, value=tpe.type_tpe.config_reseau.passerelle)
                
                ws.cell(row=row_idx, column=18, value=tpe.date_creation)
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(nom_fichier)
            return True
            
        except Exception as e:
            return False
    
    def sauvegarder(self, nom_fichier: str = None) -> bool:
        """Sauvegarde les données en format binaire (pickle)"""
        try:
            fichier = nom_fichier or self.fichier_sauvegarde
            data = {
                'tpes': [tpe.to_dict() for tpe in self.tpes],
                'date_sauvegarde': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'version': '1.5'
            }
            
            with open(fichier, 'wb') as f:
                pickle.dump(data, f)
            
            return True
            
        except Exception as e:
            return False
    
    def restaurer(self, nom_fichier: str = None) -> bool:
        """Restaure les données depuis un fichier de sauvegarde"""
        try:
            fichier = nom_fichier or self.fichier_sauvegarde
            
            if not Path(fichier).exists():
                return False
            
            with open(fichier, 'rb') as f:
                data = pickle.load(f)
            
            self.tpes = [TPE.from_dict(tpe_dict) for tpe_dict in data['tpes']]
            
            return True
            
        except Exception as e:
            return False
    
    def backup_json(self, nom_fichier: str = None) -> bool:
        """Crée une sauvegarde en format JSON (lisible)"""
        try:
            fichier = nom_fichier or self.fichier_backup
            data = {
                'tpes': [tpe.to_dict() for tpe in self.tpes],
                'date_backup': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'version': '1.5',
                'nombre_tpes': len(self.tpes)
            }
            
            with open(fichier, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            return False
    
    def restaurer_json(self, nom_fichier: str = None) -> bool:
        """Restaure les données depuis un fichier JSON"""
        try:
            fichier = nom_fichier or self.fichier_backup
            
            if not Path(fichier).exists():
                return False
            
            with open(fichier, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.tpes = [TPE.from_dict(tpe_dict) for tpe_dict in data['tpes']]
            
            return True
            
        except Exception as e:
            return False
    
    def lister_tpes(self) -> List[TPE]:
        """Retourne la liste complète des TPE"""
        return self.tpes
    
    def statistiques(self) -> dict:
        """Retourne des statistiques sur les TPE"""
        total = len(self.tpes)
        total_appareils = sum(tpe.nombre_tpe for tpe in self.tpes)
        ethernet = sum(1 for t in self.tpes if t.type_tpe.ethernet)
        quatre_cinq_g = sum(1 for t in self.tpes if t.type_tpe.quatre_cinq_g)
        backoffice_actif = sum(1 for t in self.tpes if t.acces_backoffice.actif)
        
        return {
            'total_tpes': total,
            'total_appareils': total_appareils,
            'type_ethernet': ethernet,
            'type_4_5g': quatre_cinq_g,
            'backoffice_actifs': backoffice_actif
        }